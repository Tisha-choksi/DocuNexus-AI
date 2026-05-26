import csv
import tempfile
from pathlib import Path


def _stringify(value) -> str:
    if value is None:
        return ""
    return str(value)


def _rows_to_text(rows: list[list[str]], max_rows: int = 300) -> str:
    lines = []
    for row in rows[:max_rows]:
        cleaned = [_stringify(cell).strip() for cell in row]
        if any(cleaned):
            lines.append(" | ".join(cleaned))
    if len(rows) > max_rows:
        lines.append(f"... {len(rows) - max_rows} more rows")
    return "\n".join(lines)


def _read_csv(path: Path) -> list[list[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [row for row in csv.reader(handle)]


def _read_xlsx(path: Path) -> list[tuple[str, list[list[str]]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        rows = [[_stringify(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
        sheets.append((worksheet.title, rows))
    workbook.close()
    return sheets


def _read_xls(path: Path) -> list[tuple[str, list[list[str]]]]:
    import xlrd

    workbook = xlrd.open_workbook(path)
    sheets = []
    for worksheet in workbook.sheets():
        rows = [[_stringify(worksheet.cell_value(row_index, col_index)) for col_index in range(worksheet.ncols)]
                for row_index in range(worksheet.nrows)]
        sheets.append((worksheet.name, rows))
    return sheets


def read_spreadsheet(path: Path, extension: str) -> list[tuple[str, list[list[str]]]]:
    if extension == ".csv":
        return [("CSV", _read_csv(path))]
    if extension == ".xlsx":
        return _read_xlsx(path)
    if extension == ".xls":
        return _read_xls(path)
    raise ValueError("Unsupported spreadsheet type")


def extract_spreadsheet(path: Path, extension: str) -> tuple[list[dict], dict[str, str]]:
    sheets = read_spreadsheet(path, extension)
    pages = []
    metadata = {
        "sheet_count": str(len(sheets)),
        "conversion_targets": "csv, pdf",
    }

    total_rows = 0
    for index, (sheet_name, rows) in enumerate(sheets, start=1):
        total_rows += len(rows)
        pages.append({
            "page_number": index,
            "text": f"Sheet: {sheet_name}\n{_rows_to_text(rows)}",
        })

    metadata["row_count"] = str(total_rows)
    return pages, metadata


def _first_sheet_rows(path: Path, extension: str) -> list[list[str]]:
    sheets = read_spreadsheet(path, extension)
    return sheets[0][1] if sheets else []


def spreadsheet_to_csv(path: Path, extension: str) -> Path:
    rows = _first_sheet_rows(path, extension)
    output = Path(tempfile.NamedTemporaryFile(delete=False, suffix=".csv").name)
    with output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerows(rows)
    return output


def spreadsheet_to_pdf(path: Path, extension: str, title: str) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    rows = _first_sheet_rows(path, extension)
    output = Path(tempfile.NamedTemporaryFile(delete=False, suffix=".pdf").name)
    doc = SimpleDocTemplate(str(output), pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    trimmed = rows[:80]
    if not trimmed:
        story.append(Paragraph("No spreadsheet rows were found.", styles["BodyText"]))
    else:
        max_cols = min(max(len(row) for row in trimmed), 8)
        table_data = []
        for row in trimmed:
            cells = [_stringify(cell)[:120] for cell in row[:max_cols]]
            cells.extend([""] * (max_cols - len(cells)))
            table_data.append(cells)

        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#146C94")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9E0E8")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
        if len(rows) > len(trimmed):
            story.append(Spacer(1, 10))
            story.append(Paragraph(f"Showing first {len(trimmed)} rows of {len(rows)}.", styles["BodyText"]))

    doc.build(story)
    return output


def spreadsheet_to_xlsx(path: Path, extension: str) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    sheets = read_spreadsheet(path, extension)
    output = Path(tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name)
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_names = set()

    for sheet_name, rows in sheets:
        base_name = (sheet_name or "Sheet")[:28]
        safe_name = base_name
        suffix = 1
        while safe_name in used_names:
            suffix += 1
            safe_name = f"{base_name[:25]}_{suffix}"
        used_names.add(safe_name)
        worksheet = workbook.create_sheet(safe_name)
        for row in rows:
            worksheet.append(row)
        if rows:
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="146C94")
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            worksheet.column_dimensions[column_letter].width = min(
                max(len(str(cell.value or "")) for cell in column) + 2,
                60,
            )

    if not workbook.worksheets:
        workbook.create_sheet("Sheet1")

    workbook.save(output)
    return output
